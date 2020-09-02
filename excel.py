# https://openpyxl.readthedocs.io/en/latest/
import openpyxl
from openpyxl.styles import Font, colors, Alignment
# 创建工作薄
wb = openpyxl.Workbook()

# 获取 当前工作簿处于激活状态的工作表
ws = wb.active
# Worksheel 类定义抽象"工作表"对象

# .title 属性描述了工作表的标题文字属性
# ws.title

ws.title = '成绩'
# print(ws)
# print(ws.title)
# 使用字符行列索引定位当前工作表 A1

ws['A2'] = '编号'
# 使用 cell()方法指定行列位单元格
ws.cell(row=2, column=2, value='姓名')
ws['C2'] = '成绩'

data = [
    (1, 'Tom', 90),
    (2, 'Jom', 92),
    (3, 'Som', 94),
    (4, 'Kom', 96),
    (5, 'Wom', 51)
]

# ws.append(row)
for i, row in enumerate(data):
    for j, col in enumerate(row):
        ws.cell(row=i+3, column=j+1, value=col)

ws['A1'] = '班级成绩'
ws.merge_cells('A1:C1')  # 合并一个矩形区域中的单元格
# 设置A1中的数据垂直居中和水平居中
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
# 指定字体相关，颜色，和对齐方式
bold_itatic_24_font = Font(
    name='等线', size=14, italic=True, color=colors.RED, bold=True)
ws['A1'].font = bold_itatic_24_font

for row in ws.rows:
    print(row)
# 写入保存到磁盘
wb.save('scores.xlsx')

# excel表格数据读取

my_wb = openpyxl.load_workbook(
    'scores1.xlsx', data_only=True)  # data_only = False 默认得到公式

# 查看当前工作薄下所有工作表
print(my_wb.sheetnames)
print(my_wb.active)
my_ws = my_wb['成绩']  # 找到工作表  或 my_wb.active
print(my_wb.active)  # 活动表不变

my_rows = list(my_ws.rows)

for row in my_rows:
    print(row[0].value, row[1].value, row[2].value)
